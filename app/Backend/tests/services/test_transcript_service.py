"""
Unit tests for transcript_service.py — pure pipeline functions (Steps 1-3, 5-7).

Covers:
  - read_transcript (Step 1)
  - filter_interviewer (Step 2)
  - _sent_tokenize (private helper)
  - split_sentences (Step 3)
  - select_top_n (Step 5)
  - generate_context (Step 6)
  - export_to_xlsx (Step 7)

run_predictions (Step 4) is async and depends on the NLP Docker service;
it is tested separately in integration tests.
"""

from io import BytesIO

import pandas as pd
import pytest

from transcript_service import (
    MODEL_TO_OUTCOME,
    OUTCOME_TO_SHEET,
    PHYSICIAN_IDS,
    _sent_tokenize,
    export_to_xlsx,
    filter_interviewer,
    generate_context,
    read_transcript,
    select_top_n,
    split_sentences,
)


# ── Helpers ──────────────────────────────────────────────────────────────

def _make_xlsx_bytes(df: pd.DataFrame) -> bytes:
    """Write a DataFrame to xlsx bytes (openpyxl)."""
    buf = BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    return buf.getvalue()


def _make_predicted_df(rows: int = 10) -> pd.DataFrame:
    """Build a DataFrame that resembles the output of run_predictions().

    Contains columns: index, i, i2, speaker, text, plus one column per
    outcome (cancer_prognosis, continence, ...).
    """
    data = {
        "index": list(range(1, rows + 1)),
        "i": list(range(1, rows + 1)),
        "i2": [1] * rows,
        "speaker": ["INTERVIEWER"] * rows,
        "text": [f"sentence number {i}." for i in range(1, rows + 1)],
    }
    for outcome in MODEL_TO_OUTCOME.values():
        # Scores descend so row 1 has the highest score
        data[outcome] = [round(1.0 - (j * 0.08), 6) for j in range(rows)]

    return pd.DataFrame(data)


# =========================================================================
# TestReadTranscript  (~8 tests)
# =========================================================================

class TestReadTranscript:
    """Step 1: read_transcript(file_bytes, filename)."""

    def test_reads_valid_xlsx(self):
        """Basic round-trip: valid xlsx produces a DataFrame."""
        df_in = pd.DataFrame({"speaker": ["Q", "PATIENT"], "text": ["hello", "hi"]})
        fb = _make_xlsx_bytes(df_in)
        df_out, pid = read_transcript(fb, "processed_transcripts_sid-01.xlsx")
        assert len(df_out) == 2
        assert list(df_out.columns) == ["index", "speaker", "text"]

    def test_extracts_patient_id_standard(self):
        """Standard filename format: 'processed_transcripts_sid-01.xlsx'."""
        df_in = pd.DataFrame({"speaker": ["Q"], "text": ["test"]})
        fb = _make_xlsx_bytes(df_in)
        _, pid = read_transcript(fb, "processed_transcripts_sid-01.xlsx")
        assert pid == "sid-01"

    def test_extracts_patient_id_with_prefix_stripped(self):
        """Filename with no 'processed_transcripts_' prefix keeps full stem."""
        df_in = pd.DataFrame({"speaker": ["Q"], "text": ["test"]})
        fb = _make_xlsx_bytes(df_in)
        _, pid = read_transcript(fb, "my_file.xlsx")
        assert pid == "my_file"

    def test_extracts_patient_id_complex(self):
        """Complex patient id: 'processed_transcripts_REC001 (SID 14).xlsx'."""
        df_in = pd.DataFrame({"speaker": ["Q"], "text": ["test"]})
        fb = _make_xlsx_bytes(df_in)
        _, pid = read_transcript(fb, "processed_transcripts_REC001 (SID 14).xlsx")
        assert pid == "REC001 (SID 14)"

    def test_adds_one_based_index(self):
        """Index column starts at 1 and increments."""
        df_in = pd.DataFrame({"speaker": ["Q", "Q1", "Q2"], "text": ["a", "b", "c"]})
        fb = _make_xlsx_bytes(df_in)
        df_out, _ = read_transcript(fb, "processed_transcripts_sid-99.xlsx")
        assert list(df_out["index"]) == [1, 2, 3]

    def test_raises_on_missing_columns(self):
        """ValueError when required 'speaker' or 'text' columns are absent."""
        df_in = pd.DataFrame({"col_a": [1], "col_b": [2]})
        fb = _make_xlsx_bytes(df_in)
        with pytest.raises(ValueError, match="xlsx must have columns"):
            read_transcript(fb, "test.xlsx")

    def test_raises_on_missing_text_column(self):
        """ValueError when only 'speaker' exists (no 'text')."""
        df_in = pd.DataFrame({"speaker": ["Q"], "other": ["data"]})
        fb = _make_xlsx_bytes(df_in)
        with pytest.raises(ValueError):
            read_transcript(fb, "test.xlsx")

    def test_returns_only_speaker_text_columns(self):
        """Extra columns in the source xlsx are dropped."""
        df_in = pd.DataFrame({
            "speaker": ["Q"],
            "text": ["hello"],
            "extra": [42],
        })
        fb = _make_xlsx_bytes(df_in)
        df_out, _ = read_transcript(fb, "test.xlsx")
        assert "extra" not in df_out.columns
        assert list(df_out.columns) == ["index", "speaker", "text"]


# =========================================================================
# TestFilterInterviewer  (~6 tests)
# =========================================================================

class TestFilterInterviewer:
    """Step 2: filter_interviewer(df)."""

    def _build_df(self, speakers):
        """Helper: create an indexed DataFrame with given speakers."""
        df = pd.DataFrame({
            "index": range(1, len(speakers) + 1),
            "speaker": speakers,
            "text": [f"text {i}" for i in range(len(speakers))],
        })
        return df

    def test_keeps_interviewer_speakers(self):
        """Only PHYSICIAN_IDS rows survive."""
        df = self._build_df(["INTERVIEWER", "PATIENT", "Q", "DOCTOR"])
        result = filter_interviewer(df)
        assert len(result) == 2
        assert set(result["speaker"]) == {"INTERVIEWER", "Q"}

    def test_keeps_all_physician_ids(self):
        """Every ID in PHYSICIAN_IDS is accepted."""
        df = self._build_df(PHYSICIAN_IDS)
        result = filter_interviewer(df)
        assert len(result) == len(PHYSICIAN_IDS)

    def test_removes_patient_speakers(self):
        """Non-physician speakers are removed."""
        df = self._build_df(["PATIENT", "A", "SPEAKER 1", "Doctor"])
        result = filter_interviewer(df)
        assert len(result) == 0

    def test_empty_result(self):
        """No matching speakers produces an empty DataFrame."""
        df = self._build_df(["NURSE", "AIDE"])
        result = filter_interviewer(df)
        assert len(result) == 0
        assert list(result.columns) == ["index", "speaker", "text"]

    def test_re_indexes_after_filter(self):
        """Index column is re-assigned 1..N after filtering."""
        df = self._build_df(["PATIENT", "INTERVIEWER", "PATIENT", "Q", "Q1"])
        result = filter_interviewer(df)
        assert list(result["index"]) == [1, 2, 3]

    def test_case_sensitive_matching(self):
        """Filtering is case-sensitive: 'interviewer' (lowercase) is excluded."""
        df = self._build_df(["interviewer", "INTERVIEWER", "q", "Q"])
        result = filter_interviewer(df)
        assert len(result) == 2
        assert set(result["speaker"]) == {"INTERVIEWER", "Q"}


# =========================================================================
# TestSentTokenize  (~5 tests)
# =========================================================================

class TestSentTokenize:
    """Private helper: _sent_tokenize(text)."""

    def test_splits_on_period_space(self):
        """Period followed by space triggers a split."""
        result = _sent_tokenize("First sentence. Second sentence.")
        assert result == ["First sentence.", "Second sentence."]

    def test_splits_on_exclamation(self):
        """Exclamation mark followed by space triggers a split."""
        result = _sent_tokenize("Wow! That is great.")
        assert result == ["Wow!", "That is great."]

    def test_splits_on_question_mark(self):
        """Question mark followed by space triggers a split."""
        result = _sent_tokenize("Is it good? Yes it is.")
        assert result == ["Is it good?", "Yes it is."]

    def test_single_sentence(self):
        """Single sentence without trailing space returns one element."""
        result = _sent_tokenize("Just one sentence.")
        assert result == ["Just one sentence."]

    def test_empty_string(self):
        """Empty input returns empty list."""
        result = _sent_tokenize("")
        assert result == []

    def test_whitespace_only(self):
        """Whitespace-only input returns empty list."""
        result = _sent_tokenize("   ")
        assert result == []


# =========================================================================
# TestSplitSentences  (~6 tests)
# =========================================================================

class TestSplitSentences:
    """Step 3: split_sentences(df)."""

    def _build_filtered_df(self, texts, speakers=None):
        """Helper: build a DataFrame as output of filter_interviewer."""
        n = len(texts)
        if speakers is None:
            speakers = ["INTERVIEWER"] * n
        return pd.DataFrame({
            "index": range(1, n + 1),
            "speaker": speakers,
            "text": texts,
        })

    def test_creates_correct_columns(self):
        """Output has columns: index, i, i2, speaker, text."""
        df = self._build_filtered_df(["Hello world."])
        result = split_sentences(df)
        assert list(result.columns) == ["index", "i", "i2", "speaker", "text"]

    def test_lowercases_text(self):
        """Text is lowercased (R unnest_tokens to_lower=TRUE)."""
        df = self._build_filtered_df(["Hello World."])
        result = split_sentences(df)
        assert result.iloc[0]["text"] == "hello world."

    def test_assigns_correct_i_and_i2(self):
        """i = utterance number, i2 = sentence within utterance."""
        df = self._build_filtered_df(["First. Second.", "Third."])
        result = split_sentences(df)
        # Utterance 1 has 2 sentences, utterance 2 has 1
        assert list(result["i"]) == [1, 1, 2]
        assert list(result["i2"]) == [1, 2, 1]

    def test_multi_sentence_utterance(self):
        """Multi-sentence utterance splits correctly."""
        df = self._build_filtered_df(["A. B. C."])
        result = split_sentences(df)
        assert len(result) == 3
        assert list(result["text"]) == ["a.", "b.", "c."]

    def test_empty_text_rows_skipped(self):
        """Rows with empty or whitespace text are dropped."""
        df = self._build_filtered_df(["Good sentence.", "", "   "])
        result = split_sentences(df)
        assert len(result) == 1
        assert result.iloc[0]["text"] == "good sentence."

    def test_sequential_index_numbering(self):
        """Global index is sequential 1..N across all utterances."""
        df = self._build_filtered_df(["A. B.", "C. D. E."])
        result = split_sentences(df)
        assert list(result["index"]) == [1, 2, 3, 4, 5]


# =========================================================================
# TestSelectTopN  (~6 tests)
# =========================================================================

class TestSelectTopN:
    """Step 5: select_top_n(df, n)."""

    def test_returns_all_when_n_is_zero(self):
        """n=0 returns all rows sorted by score."""
        df = _make_predicted_df(rows=5)
        result = select_top_n(df, n=0)
        for outcome in MODEL_TO_OUTCOME.values():
            assert len(result[outcome]) == 5

    def test_returns_top_n(self):
        """n>0 returns at least N rows (top scores)."""
        df = _make_predicted_df(rows=10)
        result = select_top_n(df, n=3)
        for outcome in MODEL_TO_OUTCOME.values():
            # With distinct scores, exactly N rows
            assert len(result[outcome]) == 3

    def test_includes_ties(self):
        """Tied scores at the threshold boundary are all included."""
        df = _make_predicted_df(rows=5)
        # Force rows 2 and 3 to have the same score as row 1 for one outcome
        outcome = "cancer_prognosis"
        df.loc[1, outcome] = df.loc[0, outcome]
        df.loc[2, outcome] = df.loc[0, outcome]
        result = select_top_n(df, n=1)
        # All 3 tied rows should be included
        assert len(result[outcome]) >= 3

    def test_returns_dict_with_correct_keys(self):
        """Keys are the full outcome names from MODEL_TO_OUTCOME."""
        df = _make_predicted_df(rows=3)
        result = select_top_n(df, n=0)
        expected_keys = set(MODEL_TO_OUTCOME.values())
        assert set(result.keys()) == expected_keys

    def test_each_dataframe_has_pred_1_column(self):
        """Each resulting DataFrame has a '.pred_1' column."""
        df = _make_predicted_df(rows=5)
        result = select_top_n(df, n=2)
        for outcome_df in result.values():
            assert ".pred_1" in outcome_df.columns

    def test_n_larger_than_available_rows(self):
        """n > total rows returns all rows without error."""
        df = _make_predicted_df(rows=3)
        result = select_top_n(df, n=100)
        for outcome in MODEL_TO_OUTCOME.values():
            assert len(result[outcome]) == 3


# =========================================================================
# TestGenerateContext  (~5 tests)
# =========================================================================

class TestGenerateContext:
    """Step 6: generate_context(full_df, top_df, window)."""

    def _build_sentence_df(self, n: int = 10) -> pd.DataFrame:
        """Build a DataFrame resembling split_sentences output."""
        return pd.DataFrame({
            "index": range(1, n + 1),
            "i": range(1, n + 1),
            "i2": [1] * n,
            "speaker": ["INTERVIEWER"] * n,
            "text": [f"sentence {i}" for i in range(1, n + 1)],
        })

    def _build_top_df(self, indices):
        """Build a top_df with only the specified index values."""
        rows = []
        for idx in indices:
            rows.append({
                "index": idx,
                "i": idx,
                "i2": 1,
                "speaker": "INTERVIEWER",
                "text": f"sentence {idx}",
                ".pred_1": 0.9,
            })
        return pd.DataFrame(rows)

    def test_wraps_target_in_main_tags(self):
        """Target sentence is wrapped in <main>...</main>."""
        full_df = self._build_sentence_df(5)
        top_df = self._build_top_df([3])
        contexts = generate_context(full_df, top_df, window=1)
        assert len(contexts) == 1
        assert "<main>sentence 3</main>" in contexts[0]

    def test_includes_window_sentences(self):
        """Context includes sentences within ±window of the target."""
        full_df = self._build_sentence_df(10)
        top_df = self._build_top_df([5])
        contexts = generate_context(full_df, top_df, window=2)
        context = contexts[0]
        # Should contain sentences 3, 4, <main>5</main>, 6, 7
        assert "sentence 3" in context
        assert "sentence 4" in context
        assert "<main>sentence 5</main>" in context
        assert "sentence 6" in context
        assert "sentence 7" in context

    def test_boundary_first_sentence(self):
        """Window at the start of the DataFrame clips correctly."""
        full_df = self._build_sentence_df(5)
        top_df = self._build_top_df([1])
        contexts = generate_context(full_df, top_df, window=2)
        context = contexts[0]
        # Only sentences 1, 2, 3 should be present (nothing before index 1)
        assert "<main>sentence 1</main>" in context
        assert "sentence 2" in context
        assert "sentence 3" in context
        # No negative-index sentences
        parts = context.split(".")
        assert len(parts) == 3

    def test_boundary_last_sentence(self):
        """Window at the end of the DataFrame clips correctly."""
        full_df = self._build_sentence_df(5)
        top_df = self._build_top_df([5])
        contexts = generate_context(full_df, top_df, window=2)
        context = contexts[0]
        assert "sentence 3" in context
        assert "sentence 4" in context
        assert "<main>sentence 5</main>" in context
        parts = context.split(".")
        assert len(parts) == 3

    def test_window_size_zero(self):
        """window=0 returns only the target sentence in <main> tags."""
        full_df = self._build_sentence_df(5)
        top_df = self._build_top_df([3])
        contexts = generate_context(full_df, top_df, window=0)
        assert contexts[0] == "<main>sentence 3</main>"


# =========================================================================
# TestExportToXlsx  (~4 tests)
# =========================================================================

class TestExportToXlsx:
    """Step 7: export_to_xlsx(results, patient_id)."""

    def _build_results(self) -> dict:
        """Build a results dict matching what select_top_n + context produces."""
        results = {}
        for outcome in OUTCOME_TO_SHEET:
            results[outcome] = pd.DataFrame({
                "index": [1, 2],
                "i": [1, 2],
                "i2": [1, 1],
                "speaker": ["Q", "Q"],
                "text": ["hello.", "world."],
                ".pred_1": [0.95, 0.80],
                "context": ["<main>hello.</main>", "<main>world.</main>"],
            })
        return results

    def test_creates_valid_xlsx_bytes(self):
        """Output is non-empty bytes loadable by openpyxl."""
        results = self._build_results()
        xlsx_bytes = export_to_xlsx(results, "sid-01")
        assert isinstance(xlsx_bytes, bytes)
        assert len(xlsx_bytes) > 0
        # Verify it is a valid xlsx (openpyxl can open it)
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(xlsx_bytes))
        wb.close()

    def test_has_correct_sheet_names(self):
        """Workbook contains sheets named cp, inc, ed, ius, le."""
        results = self._build_results()
        xlsx_bytes = export_to_xlsx(results, "sid-01")
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(xlsx_bytes))
        expected = set(OUTCOME_TO_SHEET.values())
        assert set(wb.sheetnames) == expected
        wb.close()

    def test_includes_name_column(self):
        """Each sheet's first column is 'name' containing the patient_id."""
        results = self._build_results()
        xlsx_bytes = export_to_xlsx(results, "sid-42")
        sheets = pd.read_excel(BytesIO(xlsx_bytes), sheet_name=None, engine="openpyxl")
        for sheet_name, df in sheets.items():
            assert "name" in df.columns, f"Sheet '{sheet_name}' missing 'name' column"
            assert (df["name"] == "sid-42").all()

    def test_readable_by_pandas(self):
        """Each sheet can be read back by pandas with correct row count."""
        results = self._build_results()
        xlsx_bytes = export_to_xlsx(results, "sid-01")
        sheets = pd.read_excel(BytesIO(xlsx_bytes), sheet_name=None, engine="openpyxl")
        for sheet_name, df in sheets.items():
            assert len(df) == 2, f"Sheet '{sheet_name}' expected 2 rows, got {len(df)}"
